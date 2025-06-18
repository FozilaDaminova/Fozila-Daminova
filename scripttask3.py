from openpyxl import load_workbook

# Load workbook and worksheet
wb = load_workbook(r"C:\Users\user\Desktop\pythonnn\sagatave_eksamenam.xlsx", data_only=True)
ws = wb['Lapa_0']

max_row = ws.max_row
header_row_number = 3

# Read headers
header = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws[header_row_number]]

# Find column indexes
adrese_col = [i for i, col in enumerate(header) if 'Adrese' in str(col)][0]
pilseta_col = [i for i, col in enumerate(header) if 'Pilsēta' in str(col)][0]

# Initialize counter
count = 0

# Loop through data
for row in range(header_row_number + 1, max_row + 1):
    adrese = ws.cell(row=row, column=adrese_col + 1).value
    pilseta = ws.cell(row=row, column=pilseta_col + 1).value

    if isinstance(adrese, str) and "Adulienas iela" in adrese:
        if pilseta in ("Valmiera", "Saulkrasti"):
            count += 1

print(f"Ierakstu skaits, kur Adrese satur 'Adulienas iela' un Pilsēta ir Valmiera vai Saulkrasti: {count}")
