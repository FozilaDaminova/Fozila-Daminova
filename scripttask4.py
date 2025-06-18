from openpyxl import load_workbook

# Load workbook and worksheet
wb = load_workbook(r"C:\Users\user\Desktop\pythonnn\sagatave_eksamenam.xlsx", data_only=True)
ws = wb['Lapa_0']

max_row = ws.max_row
header_row_number = 3

# Read headers
header = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws[header_row_number]]

# Find column indexes
produkts_col = [i for i, col in enumerate(header) if 'Produkts' in str(col)][0]
cena_col = [i for i, col in enumerate(header) if 'Cena' in str(col)][0]

# Initialize sum and count
total_sum = 0
count = 0

# Loop through rows
for row in range(header_row_number + 1, max_row + 1):
    produkts = ws.cell(row=row, column=produkts_col + 1).value
    cena = ws.cell(row=row, column=cena_col + 1).value

    if isinstance(produkts, str) and "LaserJet" in produkts:
        try:
            total_sum += float(cena)
            count += 1
        except:
            continue

# Calculate average and round down
if count > 0:
    average = int(total_sum / count)
    print(f"Average Cena for products containing 'LaserJet': {average}")
else:
    print("No matching LaserJet products found.")
