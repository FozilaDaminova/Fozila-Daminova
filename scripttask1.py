from openpyxl import load_workbook

# Load workbook and worksheet
wb = load_workbook(r"C:\Users\user\Desktop\pythonnn\sagatave_eksamenam.xlsx", data_only=True)
ws = wb['Lapa_0']

max_row = ws.max_row
header_row_number = 3

# Get headers
header = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws[header_row_number]]

# Get column indexes
adrese_col = [i for i, col in enumerate(header) if 'Adrese' in str(col)][0]
skaits_col = [i for i, col in enumerate(header) if 'Skaits' in str(col)][0]

# Count matching entries
count = 0
for row in range(header_row_number + 1, max_row + 1):
    adrese = ws.cell(row=row, column=adrese_col + 1).value
    skaits = ws.cell(row=row, column=skaits_col + 1).value

    if isinstance(adrese, str) and adrese.startswith("Ain"):
        try:
            if float(skaits) < 40:
                count += 1
        except:
            continue

print(f"Ierakstu skaits, kur Adrese sākas ar 'Ain' un Skaits < 40: {count}")



