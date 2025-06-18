from openpyxl import load_workbook

# Load workbook and worksheet
wb = load_workbook(r"C:\Users\user\Desktop\pythonnn\sagatave_eksamenam.xlsx", data_only=True)
ws = wb['Lapa_0']

max_row = ws.max_row
header_row_number = 3

# Get headers
header = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws[header_row_number]]

# Find column indexes
klients_col = [i for i, col in enumerate(header) if 'Klients' in str(col)][0]
skaits_col = [i for i, col in enumerate(header) if 'Skaits' in str(col)][0]
kopa_col = [i for i, col in enumerate(header) if 'Kopā' in str(col)][0]

# Initialize total
total_sum = 0

# Loop through data rows
for row in range(header_row_number + 1, max_row + 1):
    klients = ws.cell(row=row, column=klients_col + 1).value
    skaits = ws.cell(row=row, column=skaits_col + 1).value
    kopa = ws.cell(row=row, column=kopa_col + 1).value

    try:
        if klients == "Korporatīvais" and 40 <= float(skaits) <= 50:
            total_sum += float(kopa)
    except:
        continue

# Round down total
print(f"Kopā summa (Korporatīvais klienti, Skaits 40–50): {int(total_sum)}")
