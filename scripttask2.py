from openpyxl import load_workbook
from datetime import datetime

# Load workbook and worksheet
wb = load_workbook(r"C:\Users\user\Desktop\pythonnn\sagatave_eksamenam.xlsx", data_only=True)
ws = wb['Lapa_0']

max_row = ws.max_row

# ✅ Read headers from row 3 (instead of row 2)
header_row_number = 3
header = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws[header_row_number]]

# Find the correct column indices
prioritate_col = [i for i, col in enumerate(header) if 'Prioritāte' in str(col)][0]
piegades_datums_col = [i for i, col in enumerate(header) if 'Piegādes datums' in str(col)][0]

count = 0

# ✅ Start data from the row after the header
for row in range(header_row_number + 1, max_row + 1):
    prioritāte = ws.cell(row=row, column=prioritate_col + 1).value
    datums = ws.cell(row=row, column=piegades_datums_col + 1).value

    if prioritāte == "High":
        try:
            year = datums.year if isinstance(datums, datetime) else datetime.strptime(str(datums), "%Y-%m-%d %H:%M:%S").year
            if year == 2015:
                count += 1
        except:
            continue

print(f"Number of entries with High priority and delivery year 2015: {count}")